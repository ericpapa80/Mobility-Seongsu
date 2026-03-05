#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TOPIS 정적 참조 데이터 → 성수 권역 링크 JSON 생성

입력:
  - pipeline/source/topis/서비스링크 보간점 정보(LINK_VERTEX)_2025.xlsx  (좌표)
  - pipeline/source/topis/서울시 도로 기능별 구분 정보_2025.xlsx          (도로속성)

출력:
  - pipeline/ref/topis_seongsu_links.json

좌표: GRS80TM (EPSG:5181) → WGS84 (EPSG:4326) 변환
"""

import json
import sys
import io
from pathlib import Path

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from pyproj import Transformer
except ImportError:
    print("필요 패키지: pip install openpyxl pyproj")
    sys.exit(1)

PIPELINE = Path(__file__).resolve().parent.parent
SOURCE_DIR = PIPELINE / "source" / "topis"
REF_DIR = PIPELINE / "ref"

BBOX_MIN_LNG, BBOX_MIN_LAT = 127.025, 37.525
BBOX_MAX_LNG, BBOX_MAX_LAT = 127.080, 37.565


def parse_link_vertices() -> dict[str, list[tuple[float, float]]]:
    """서비스링크 보간점 XLSX → LINK_ID별 WGS84 좌표 리스트"""
    src = SOURCE_DIR / "서비스링크 보간점 정보(LINK_VERTEX)_2025.xlsx"
    print(f"1. 서비스링크 보간점 로딩: {src.name}")

    transformer = Transformer.from_crs("EPSG:5181", "EPSG:4326", always_xy=True)

    wb = openpyxl.load_workbook(src, read_only=True)
    ws = wb.active

    links: dict[str, list[tuple[int, float, float]]] = {}
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        lid = str(row[0]) if row[0] else None
        if not lid or row[2] is None or row[3] is None:
            skipped += 1
            continue

        seq = row[1] or 0
        x, y = float(row[2]), float(row[3])
        lng, lat = transformer.transform(x, y)
        if lid not in links:
            links[lid] = []
        links[lid].append((seq, round(lng, 6), round(lat, 6)))

    wb.close()
    print(f"   → 전체 {len(links)}개 링크, {skipped}개 빈 행 스킵")

    result: dict[str, list[tuple[float, float]]] = {}
    for lid, verts in links.items():
        verts.sort(key=lambda v: v[0])
        result[lid] = [(lng, lat) for _, lng, lat in verts]

    return result


def parse_road_info() -> dict[int, dict]:
    """도로 기능별 구분 정보 XLSX → 도로축코드별 속성"""
    src = SOURCE_DIR / "서울시 도로 기능별 구분 정보_2025.xlsx"
    print(f"2. 도로 기능별 구분 정보 로딩: {src.name}")

    wb = openpyxl.load_workbook(src, read_only=True)
    ws = wb.active

    roads: dict[int, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        code = int(row[0])
        roads[code] = {
            "road_name": row[1] or "",
            "road_type_code": int(row[2]) if row[2] else 0,
            "road_type": row[3] or "",
        }

    wb.close()
    print(f"   → {len(roads)}개 도로 정보")
    return roads


def filter_by_bbox(
    all_links: dict[str, list[tuple[float, float]]],
) -> dict[str, list[tuple[float, float]]]:
    """확장 bbox 내 최소 1개 vertex가 있는 링크만 필터"""
    filtered = {}
    for lid, coords in all_links.items():
        for lng, lat in coords:
            if (BBOX_MIN_LNG <= lng <= BBOX_MAX_LNG
                    and BBOX_MIN_LAT <= lat <= BBOX_MAX_LAT):
                filtered[lid] = coords
                break
    return filtered


def load_geojson_road_attrs() -> dict[str, dict]:
    """기존 ss_pl_traffic.geojson에서 LINK_ID별 도로 속성 추출"""
    geo_path = REF_DIR / "ss_pl_traffic.geojson"
    if not geo_path.exists():
        print("   ⚠ ss_pl_traffic.geojson 없음 — 도로명 매핑 건너뜀")
        return {}

    print(f"3. 도로 속성 매핑 로딩: {geo_path.name}")
    with open(geo_path, encoding="utf-8") as f:
        geo = json.load(f)

    attrs: dict[str, dict] = {}
    for feat in geo["features"]:
        props = feat["properties"]
        lid = str(props.get("LINK_ID", ""))
        attrs[lid] = {
            "road_name": props.get("도로명", ""),
            "direction": props.get("방향", ""),
            "distance": props.get("거리", 0),
            "lanes": props.get("차선수", 1),
            "road_type": props.get("기능유형구분", ""),
        }
    print(f"   → {len(attrs)}개 링크 속성")
    return attrs


def main():
    all_links = parse_link_vertices()
    road_info = parse_road_info()
    geo_attrs = load_geojson_road_attrs()

    filtered = filter_by_bbox(all_links)
    print(f"4. 확장 bbox 필터: {len(filtered)}/{len(all_links)}개 링크")
    print(f"   bbox: [{BBOX_MIN_LNG}, {BBOX_MIN_LAT}, {BBOX_MAX_LNG}, {BBOX_MAX_LAT}]")

    link_ids = sorted(filtered.keys())
    matched = sum(1 for lid in link_ids if lid in geo_attrs)
    print(f"   도로명 매칭: {matched}/{len(link_ids)}개")

    links_out = []
    for lid in link_ids:
        coords = filtered[lid]
        attr = geo_attrs.get(lid, {})
        links_out.append({
            "link_id": lid,
            "road_name": attr.get("road_name", ""),
            "direction": attr.get("direction", ""),
            "distance": attr.get("distance", 0),
            "lanes": attr.get("lanes", 1),
            "road_type": attr.get("road_type", ""),
            "coordinates": [list(c) for c in coords],
        })

    output = {
        "meta": {
            "bbox": [BBOX_MIN_LNG, BBOX_MIN_LAT, BBOX_MAX_LNG, BBOX_MAX_LAT],
            "link_count": len(links_out),
            "source": "TOPIS 서비스링크 보간점 정보 2025",
            "crs_from": "EPSG:5181",
            "crs_to": "EPSG:4326",
        },
        "link_ids": link_ids,
        "road_types": {
            str(code): info for code, info in road_info.items()
        },
        "links": links_out,
    }

    REF_DIR.mkdir(parents=True, exist_ok=True)
    out_path = REF_DIR / "topis_seongsu_links.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False)

    print(f"4. 출력 완료: {out_path.name} ({len(links_out)}개 링크)")


if __name__ == "__main__":
    main()
